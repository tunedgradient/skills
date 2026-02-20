#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sys
from calendar import monthrange
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

SCALE_MULTIPLIERS = {
    "ones": 1.0,
    "usd": 1.0,
    "usd_000": 1_000.0,
    "usd_mm": 1_000_000.0,
    "usd_bn": 1_000_000_000.0,
    "eur": 1.0,
    "eur_000": 1_000.0,
    "eur_mm": 1_000_000.0,
    "eur_bn": 1_000_000_000.0,
    "gbp": 1.0,
    "gbp_000": 1_000.0,
    "gbp_mm": 1_000_000.0,
    "gbp_bn": 1_000_000_000.0,
    "jpy": 1.0,
    "jpy_000": 1_000.0,
    "jpy_mm": 1_000_000.0,
    "jpy_bn": 1_000_000_000.0,
}

OUTPUT_FIELDS = [
    "entity",
    "metric_raw",
    "metric",
    "value_raw",
    "value",
    "unit",
    "scale",
    "window_label",
    "window_end_date",
    "window_start_date",
    "as_of_date",
    "source",
    "notes",
]

METRIC_COL_CANDIDATES = (
    "metric",
    "line item",
    "line_item",
    "kpi",
    "account",
    "description",
    "item",
    "name",
)
VALUE_COL_CANDIDATES = ("value", "amount", "reported value", "reported_value")
WINDOW_COL_CANDIDATES = (
    "window",
    "period",
    "reporting window",
    "reporting_window",
    "time period",
    "time_period",
    "quarter",
    "fiscal period",
    "fiscal_period",
)
UNIT_COL_CANDIDATES = (
    "unit",
    "currency",
    "ccy",
    "unit_hint",
    "currency_code",
    "scale",
)
AS_OF_COL_CANDIDATES = ("as_of_date", "as of", "as_of", "filing date", "report date")

CURRENCY_HINTS = [
    ("USD", [r"\bUSD\b", r"US\$", r"\$"]),
    ("EUR", [r"\bEUR\b", r"€"]),
    ("GBP", [r"\bGBP\b", r"£"]),
    ("JPY", [r"\bJPY\b", r"¥"]),
]

SCALE_HINTS = [
    ("_bn", [r"\bbn\b", r"\bbillion(s)?\b", r"\bin\s+billions\b"]),
    ("_mm", [r"\bmm\b", r"\bmillion(s)?\b", r"\bin\s+millions\b"]),
    ("_000", [r"\b000s\b", r"\bthousand(s)?\b", r"\bin\s+thousands\b"]),
]

WINDOW_PATTERNS = [
    (re.compile(r"\bQ([1-4])\s*['’]?\s*(\d{2}|\d{4})\b", re.IGNORECASE), "quarter"),
    (re.compile(r"\bQ([1-4])\b", re.IGNORECASE), "quarter_no_year"),
    (re.compile(r"\bFY\s*['’]?\s*(\d{2}|\d{4})\b", re.IGNORECASE), "fiscal_year"),
    (re.compile(r"\b(LTM|TTM)\b", re.IGNORECASE), "rolling"),
    (re.compile(r"\bYTD\s*['’]?\s*(\d{2}|\d{4})\b", re.IGNORECASE), "ytd"),
    (re.compile(r"\bYTD\b", re.IGNORECASE), "ytd_no_year"),
    (
        re.compile(r"\b(9M|6M|3M|1H|2H)\s*['’]?\s*(\d{2}|\d{4})\b", re.IGNORECASE),
        "partial_year",
    ),
    (re.compile(r"\b(9M|6M|3M|1H|2H)\b", re.IGNORECASE), "partial_year_no_year"),
]

CURRENCY_METRIC_KEYWORDS = (
    "revenue",
    "sales",
    "income",
    "cash",
    "debt",
    "capex",
    "opex",
    "ebit",
    "ebitda",
    "asset",
    "liability",
    "expense",
    "cost",
    "profit",
    "book value",
)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Normalize units and reporting windows in finance tables.")
    ap.add_argument("--input", required=True, help="Input CSV or XLSX path.")
    ap.add_argument("--output", default="normalized_financials.csv", help="Normalized CSV output path.")
    ap.add_argument("--notes", default="normalization_notes.md", help="Normalization notes output path.")
    ap.add_argument("--sheet", default="", help="Optional XLSX sheet name.")
    ap.add_argument("--entity", default="", help="Optional entity label.")
    ap.add_argument("--source", default="", help="Optional source override.")
    return ap.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", normalize_text(name).lower()).strip("_")


def normalize_year(raw_year: str) -> str:
    y = normalize_text(raw_year)
    if len(y) == 2 and y.isdigit():
        n = int(y)
        return str(2000 + n) if n <= 79 else str(1900 + n)
    return y


def first_matching_column(fieldnames: Sequence[str], candidates: Sequence[str]) -> str:
    normalized_candidates = {normalize_header(c) for c in candidates}
    for field in fieldnames:
        if normalize_header(field) in normalized_candidates:
            return field
    return ""


def sniff_currency(text: str) -> str:
    haystack = text or ""
    for code, patterns in CURRENCY_HINTS:
        for pattern in patterns:
            if re.search(pattern, haystack):
                return code
    return ""


def sniff_scale_suffix(text: str) -> str:
    haystack = (text or "").lower()
    for suffix, patterns in SCALE_HINTS:
        for pattern in patterns:
            if re.search(pattern, haystack):
                return suffix
    return ""


def parse_window_label(text: str) -> Tuple[str, List[str]]:
    raw = normalize_text(text)
    if not raw:
        return "", []

    notes: List[str] = []
    for pattern, kind in WINDOW_PATTERNS:
        match = pattern.search(raw)
        if not match:
            continue
        if kind == "quarter":
            return f"Q{match.group(1)} {normalize_year(match.group(2))}", notes
        if kind == "quarter_no_year":
            notes.append("window_year_unknown")
            return f"Q{match.group(1)}", notes
        if kind == "fiscal_year":
            return f"FY {normalize_year(match.group(1))}", notes
        if kind == "rolling":
            return match.group(1).upper(), notes
        if kind == "ytd":
            return f"YTD {normalize_year(match.group(1))}", notes
        if kind == "ytd_no_year":
            notes.append("window_year_unknown")
            return "YTD", notes
        if kind == "partial_year":
            return f"{match.group(1).upper()} {normalize_year(match.group(2))}", notes
        if kind == "partial_year_no_year":
            notes.append("window_year_unknown")
            return match.group(1).upper(), notes

    year_only = re.fullmatch(r"(?:FY\s*)?['’]?(\d{2}|\d{4})A?", raw, flags=re.IGNORECASE)
    if year_only:
        notes.append("year_only_label_assumed_fy")
        return f"FY {normalize_year(year_only.group(1))}", notes

    return "", []


def derive_window_dates(window_label: str) -> Tuple[str, str, List[str]]:
    if not window_label:
        return "", "", []

    notes: List[str] = []
    quarter_match = re.fullmatch(r"Q([1-4]) (\d{4})", window_label)
    if quarter_match:
        quarter = int(quarter_match.group(1))
        year = int(quarter_match.group(2))
        start_month = 1 + (quarter - 1) * 3
        end_month = quarter * 3
        start_date = date(year, start_month, 1).isoformat()
        end_day = monthrange(year, end_month)[1]
        end_date = date(year, end_month, end_day).isoformat()
        return start_date, end_date, notes

    return "", "", notes


def parse_number(raw_value: Any) -> Tuple[Optional[float], str]:
    if raw_value is None:
        return None, "missing"

    if isinstance(raw_value, (int, float)):
        return float(raw_value), ""

    value = normalize_text(raw_value)
    if value == "":
        return None, "blank"

    if value.upper() in {"N/A", "NA", "NM", "N.M.", "-", "--", "—"}:
        return None, "not_numeric"

    is_percent = value.endswith("%")
    if is_percent:
        value = value[:-1].strip()

    is_negative = value.startswith("(") and value.endswith(")")
    if is_negative:
        value = value[1:-1].strip()

    value = value.replace(",", "").replace(" ", "")
    value = value.replace("$", "").replace("€", "").replace("£", "").replace("¥", "")
    value = re.sub(r"\b(USD|EUR|GBP|JPY)\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"[\*]+$", "", value)

    if value == "":
        return None, "num_parse_fail"

    try:
        parsed = float(value)
        if is_negative:
            parsed = -parsed
        if is_percent:
            return parsed / 100.0, "percent"
        return parsed, ""
    except ValueError:
        return None, "num_parse_fail"


def format_value(value: Any) -> str:
    if value == "":
        return ""
    if isinstance(value, (int, float)):
        numeric = float(value)
        if abs(numeric) < 1e-12:
            numeric = 0.0
        return f"{numeric:.10g}"
    return str(value)


def read_csv_rows(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows: List[Dict[str, Any]] = []
        for row in reader:
            rows.append({key: row.get(key, "") for key in (reader.fieldnames or [])})
        return list(reader.fieldnames or []), rows


def dedupe_headers(headers: Iterable[str]) -> List[str]:
    output: List[str] = []
    counts: Counter[str] = Counter()
    for idx, raw in enumerate(headers, start=1):
        header = normalize_text(raw) or f"column_{idx}"
        counts[header] += 1
        if counts[header] > 1:
            header = f"{header}_{counts[header]}"
        output.append(header)
    return output


def read_xlsx_rows(path: Path, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "XLSX input requires openpyxl. Install with: python3 -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise RuntimeError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    row_iter = worksheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        return [], []

    headers = dedupe_headers([normalize_text(cell) for cell in header_row])
    rows: List[Dict[str, Any]] = []
    for raw_row in row_iter:
        record: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            record[header] = raw_row[idx] if idx < len(raw_row) else ""
        rows.append(record)
    return headers, rows


def load_input(path: Path, sheet_name: str) -> Tuple[List[str], List[Dict[str, Any]], str]:
    ext = path.suffix.lower()
    if ext == ".csv":
        fields, rows = read_csv_rows(path)
        return fields, rows, "csv"
    if ext in {".xlsx", ".xlsm"}:
        fields, rows = read_xlsx_rows(path, sheet_name)
        return fields, rows, "xlsx"
    raise RuntimeError(f"Unsupported input extension '{ext}'. Use CSV or XLSX.")


def detect_shape(fieldnames: Sequence[str]) -> Tuple[str, List[str]]:
    value_col = first_matching_column(fieldnames, VALUE_COL_CANDIDATES)
    metric_col = first_matching_column(fieldnames, METRIC_COL_CANDIDATES)

    if value_col and metric_col:
        return "long", []

    window_columns = []
    for header in fieldnames:
        window_label, _ = parse_window_label(header)
        if window_label:
            window_columns.append(header)

    if window_columns:
        return "wide", window_columns

    if value_col:
        return "long", []

    return "wide", []


def infer_metric_flags(metric_raw: str, parse_note: str) -> Tuple[bool, bool]:
    metric_text = metric_raw.lower()
    percent_metric = parse_note == "percent" or "%" in metric_raw or "percent" in metric_text
    per_share_metric = bool(re.search(r"\beps\b|per[\s-]*share|/share", metric_text))
    return percent_metric, per_share_metric


def detect_currency_metric(metric_raw: str, value_raw: str) -> bool:
    merged = f"{metric_raw} {value_raw}"
    if sniff_currency(merged):
        return True

    metric_text = metric_raw.lower()
    return any(keyword in metric_text for keyword in CURRENCY_METRIC_KEYWORDS)


def scale_value(value: Optional[float], scale: str, apply_table_scaling: bool) -> Tuple[Any, List[str]]:
    if value is None:
        return "", []

    notes: List[str] = []
    if not apply_table_scaling:
        return value, notes

    multiplier = SCALE_MULTIPLIERS.get(scale)
    if multiplier is None:
        notes.append("scale_multiplier_missing")
        return value, notes
    return value * multiplier, notes


def collect_global_hints(fieldnames: Sequence[str], rows: Sequence[Dict[str, Any]]) -> Tuple[str, str]:
    header_text = " | ".join(fieldnames)
    sample_chunks = []
    for row in rows[:10]:
        sample_chunks.append(" ".join(normalize_text(v) for v in row.values()))
    combined = f"{header_text} {' '.join(sample_chunks)}"
    return sniff_currency(combined), sniff_scale_suffix(combined)


def coalesce(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def build_long_observations(
    rows: Sequence[Dict[str, Any]],
    fieldnames: Sequence[str],
) -> List[Dict[str, str]]:
    metric_col = first_matching_column(fieldnames, METRIC_COL_CANDIDATES) or (fieldnames[0] if fieldnames else "")
    value_col = first_matching_column(fieldnames, VALUE_COL_CANDIDATES)
    window_col = first_matching_column(fieldnames, WINDOW_COL_CANDIDATES)
    unit_col = first_matching_column(fieldnames, UNIT_COL_CANDIDATES)
    as_of_col = first_matching_column(fieldnames, AS_OF_COL_CANDIDATES)

    observations: List[Dict[str, str]] = []
    for row in rows:
        observations.append(
            {
                "metric_raw": coalesce(row.get(metric_col)),
                "value_raw": coalesce(row.get(value_col)),
                "window_raw": coalesce(row.get(window_col)),
                "unit_hint": coalesce(row.get(unit_col)),
                "as_of_date": coalesce(row.get(as_of_col)),
                "row_context": " | ".join(coalesce(row.get(name)) for name in fieldnames),
            }
        )
    return observations


def build_wide_observations(
    rows: Sequence[Dict[str, Any]],
    fieldnames: Sequence[str],
    explicit_window_columns: Sequence[str],
) -> List[Dict[str, str]]:
    unit_col = first_matching_column(fieldnames, UNIT_COL_CANDIDATES)
    as_of_col = first_matching_column(fieldnames, AS_OF_COL_CANDIDATES)
    metric_col = first_matching_column(fieldnames, METRIC_COL_CANDIDATES)

    window_columns: List[str] = list(explicit_window_columns)
    if not window_columns:
        for header in fieldnames:
            window_label, _ = parse_window_label(header)
            if window_label:
                window_columns.append(header)
    if not window_columns:
        raise RuntimeError(
            "Could not detect wide-table window columns (for example Q1 2025, FY 2024, LTM, YTD 2025, 9M 2025)."
        )

    excluded = set(window_columns)
    if unit_col:
        excluded.add(unit_col)
    if as_of_col:
        excluded.add(as_of_col)

    if not metric_col:
        for header in fieldnames:
            if header not in excluded:
                metric_col = header
                break

    observations: List[Dict[str, str]] = []
    for row in rows:
        row_context = " | ".join(coalesce(row.get(name)) for name in fieldnames)
        metric_raw = coalesce(row.get(metric_col))
        unit_hint = coalesce(row.get(unit_col))
        as_of_date = coalesce(row.get(as_of_col))
        for window_col in window_columns:
            value_raw = coalesce(row.get(window_col))
            if value_raw == "":
                continue
            observations.append(
                {
                    "metric_raw": metric_raw,
                    "value_raw": value_raw,
                    "window_raw": window_col,
                    "unit_hint": unit_hint,
                    "as_of_date": as_of_date,
                    "row_context": row_context,
                }
            )
    return observations


def choose_scale_and_unit(
    observation: Dict[str, str],
    global_currency: str,
    global_scale_suffix: str,
    percent_metric: bool,
    per_share_metric: bool,
) -> Tuple[str, str, bool, List[str]]:
    notes: List[str] = []

    metric_raw = observation.get("metric_raw", "")
    value_raw = observation.get("value_raw", "")
    unit_hint = observation.get("unit_hint", "")
    row_context = observation.get("row_context", "")
    merged_text = " ".join([metric_raw, value_raw, unit_hint, row_context])

    row_currency = sniff_currency(unit_hint) or sniff_currency(merged_text) or global_currency
    row_scale_suffix = sniff_scale_suffix(unit_hint) or sniff_scale_suffix(merged_text) or global_scale_suffix

    if percent_metric:
        return "", "ratio", False, ["percent_metric_no_currency_scale"]

    if per_share_metric:
        return row_currency, "per_share", False, ["per_share_no_table_scale"]

    if row_currency:
        scale = row_currency.lower() + (row_scale_suffix or "")
        if scale not in SCALE_MULTIPLIERS:
            if row_scale_suffix:
                notes.append("scale_ambiguous")
            scale = row_currency.lower()
        return row_currency, scale, True, notes

    if row_scale_suffix:
        notes.append("scale_hint_without_currency")
    return "", "ones", True, notes


def normalize_observations(
    observations: Sequence[Dict[str, str]],
    output_path: Path,
    entity: str,
    source_label: str,
    global_currency: str,
    global_scale_suffix: str,
) -> Tuple[Counter, List[str]]:
    stats: Counter = Counter()
    assumptions: List[str] = []

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()

        for observation in observations:
            stats["rows_seen"] += 1

            metric_raw = observation.get("metric_raw", "")
            value_raw = observation.get("value_raw", "")
            window_raw = observation.get("window_raw", "")
            as_of_date = observation.get("as_of_date", "")

            parsed_value, parse_note = parse_number(value_raw)
            percent_metric, per_share_metric = infer_metric_flags(metric_raw, parse_note)
            currency_metric = detect_currency_metric(metric_raw, value_raw)

            window_label, window_notes = parse_window_label(window_raw)
            if not window_label:
                inferred, inferred_notes = parse_window_label(metric_raw)
                window_label = inferred
                window_notes.extend(inferred_notes)
            if not window_label and window_raw:
                window_label = window_raw
                window_notes.append("window_unparsed")

            window_start_date, window_end_date, window_date_notes = derive_window_dates(window_label)

            unit, scale, apply_scaling, scale_notes = choose_scale_and_unit(
                observation,
                global_currency=global_currency,
                global_scale_suffix=global_scale_suffix,
                percent_metric=percent_metric,
                per_share_metric=per_share_metric,
            )

            scaled_value, scaling_notes = scale_value(parsed_value, scale, apply_table_scaling=apply_scaling)

            row_notes: List[str] = []
            if parse_note:
                row_notes.append(parse_note)
            row_notes.extend(window_notes)
            row_notes.extend(window_date_notes)
            row_notes.extend(scale_notes)
            row_notes.extend(scaling_notes)

            if not window_label:
                stats["missing_window_label"] += 1
                row_notes.append("missing_window_label")

            if parse_note == "num_parse_fail":
                stats["num_parse_fail"] += 1
            if parse_note == "percent":
                stats["percent_rows"] += 1
            if per_share_metric:
                stats["per_share_rows"] += 1

            if currency_metric and not unit and not percent_metric and not per_share_metric:
                stats["currency_ambiguous"] += 1
                row_notes.append("currency_ambiguous")

            if "scale_ambiguous" in row_notes:
                stats["scale_ambiguous"] += 1

            if "window_year_unknown" in row_notes:
                stats["window_year_unknown"] += 1

            writer.writerow(
                {
                    "entity": entity,
                    "metric_raw": metric_raw,
                    "metric": "",
                    "value_raw": value_raw,
                    "value": format_value(scaled_value),
                    "unit": unit,
                    "scale": scale,
                    "window_label": window_label,
                    "window_end_date": window_end_date,
                    "window_start_date": window_start_date,
                    "as_of_date": as_of_date,
                    "source": source_label,
                    "notes": ";".join(dict.fromkeys([n for n in row_notes if n])),
                }
            )

            stats["rows_written"] += 1

    if not global_currency:
        assumptions.append("Currency hint was not detected globally; currency may be ambiguous by row.")
    if not global_scale_suffix:
        assumptions.append("Scale hint was not detected globally; default currency scale falls back to base units.")
    if stats["missing_window_label"] > 0:
        assumptions.append("Some rows have missing or unparseable reporting windows; see row notes.")
    if stats["currency_ambiguous"] > 0:
        assumptions.append("Some currency-like metrics lacked confident currency detection; see row notes.")
    if stats["scale_ambiguous"] > 0:
        assumptions.append("Some rows had conflicting scale hints; conservative scale fallback was applied.")

    return stats, assumptions


def write_notes(
    notes_path: Path,
    input_path: Path,
    input_kind: str,
    inferred_shape: str,
    entity: str,
    source_label: str,
    global_currency: str,
    global_scale_suffix: str,
    stats: Counter,
    assumptions: Sequence[str],
) -> None:
    lines: List[str] = []
    lines.append("# Normalization notes")
    lines.append("")
    lines.append("## Run context")
    lines.append(f"- input: `{input_path}`")
    lines.append(f"- input_type: `{input_kind}`")
    lines.append(f"- inferred_shape: `{inferred_shape}`")
    lines.append(f"- entity: `{entity or 'n/a'}`")
    lines.append(f"- source: `{source_label}`")
    lines.append("- currency_conversion: `disabled` (v1 keeps values as reported)")
    lines.append("")
    lines.append("## Global detection")
    lines.append(f"- currency_hint: `{global_currency or 'unknown'}`")
    lines.append(f"- scale_hint: `{global_scale_suffix or 'unknown'}`")
    lines.append("")
    lines.append("## Row stats")
    lines.append(f"- rows_seen: `{stats['rows_seen']}`")
    lines.append(f"- rows_written: `{stats['rows_written']}`")
    lines.append(f"- num_parse_fail: `{stats['num_parse_fail']}`")
    lines.append(f"- percent_rows: `{stats['percent_rows']}`")
    lines.append(f"- per_share_rows: `{stats['per_share_rows']}`")
    lines.append(f"- missing_window_label: `{stats['missing_window_label']}`")
    lines.append(f"- window_year_unknown: `{stats['window_year_unknown']}`")
    lines.append(f"- currency_ambiguous: `{stats['currency_ambiguous']}`")
    lines.append(f"- scale_ambiguous: `{stats['scale_ambiguous']}`")
    lines.append("")
    lines.append("## Assumptions and limitations")
    if assumptions:
        for assumption in assumptions:
            lines.append(f"- {assumption}")
    else:
        lines.append("- No major assumptions were required.")
    lines.append("")
    notes_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    notes_path = Path(args.notes).expanduser().resolve()

    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    try:
        fieldnames, rows, input_kind = load_input(input_path, args.sheet)
        if not fieldnames:
            raise RuntimeError("Input has no columns.")

        inferred_shape, explicit_window_columns = detect_shape(fieldnames)
        global_currency, global_scale_suffix = collect_global_hints(fieldnames, rows)

        if inferred_shape == "long":
            observations = build_long_observations(rows, fieldnames)
        else:
            observations = build_wide_observations(rows, fieldnames, explicit_window_columns)

        source_label = args.source or input_path.name
        stats, assumptions = normalize_observations(
            observations=observations,
            output_path=output_path,
            entity=args.entity,
            source_label=source_label,
            global_currency=global_currency,
            global_scale_suffix=global_scale_suffix,
        )
        write_notes(
            notes_path=notes_path,
            input_path=input_path,
            input_kind=input_kind,
            inferred_shape=inferred_shape,
            entity=args.entity,
            source_label=source_label,
            global_currency=global_currency,
            global_scale_suffix=global_scale_suffix,
            stats=stats,
            assumptions=assumptions,
        )
    except RuntimeError as exc:
        raise SystemExit(str(exc)) from exc


if __name__ == "__main__":
    try:
        main()
    except BrokenPipeError:
        # Allow piping into commands like `head` without stack traces.
        sys.exit(0)
